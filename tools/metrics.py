#!/usr/bin/env python3
"""Compute what the system reports about itself.

Everything here is derived from the matrix and the ledger — there is no separate
tracking to maintain, and nothing is invented. The one assumption (how long it takes
you to review a call by hand) comes from config and is always printed next to the
number it produces.

Usage:
    python3 tools/metrics.py                      # print a summary
    python3 tools/metrics.py --write              # also append a row to 'Run metrics'
    python3 tools/metrics.py --scope week         # week (default) | run | all
    python3 tools/metrics.py --markdown out.md    # write a shareable summary
"""
import argparse
import collections
import datetime as dt
import os
import re
import sys

from openpyxl import load_workbook

DEFAULT_MINUTES_PER_CALL = 15  # fallback ONLY for calls whose real length is unknown
MIN_REAL_CALL_MINUTES = 5      # shorter than this is a no-show, not a conversation


def read_config(path):
    """Minimal config reader: the few scalars metrics needs, no YAML dependency."""
    cfg = {"minutes_per_call_review": DEFAULT_MINUTES_PER_CALL,
           "matrix": os.path.join("output", "signal-matrix.xlsx"),
           "company": "Your company"}
    if not os.path.exists(path):
        return cfg
    section = None
    for line in open(path, encoding="utf-8"):
        if line.strip().startswith("#") or not line.strip():
            continue
        if not line[0].isspace():
            section = line.split(":", 1)[0].strip()
        stripped = line.strip()
        if ":" not in stripped:
            continue
        key, _, raw = stripped.partition(":")
        key, raw = key.strip(), raw.strip()
        # Quoted values win outright, so a hex colour is not mistaken for a comment.
        q = re.match(r'^"([^"]*)"|^\'([^\']*)\'', raw)
        val = (q.group(1) if q.group(1) is not None else q.group(2)) if q else raw.split("#")[0].strip()
        if not val:
            continue
        if key == "minutes_per_call_review":
            cfg["minutes_per_call_review"] = int(val)
        elif key == "name" and section == "company":
            cfg["company"] = val
        elif key == "matrix_filename":
            cfg["matrix_filename"] = val
        elif key == "dir" and section == "output":
            cfg["output_dir"] = val
    # Relative paths in the config resolve against the config file's own directory.
    base = os.path.dirname(os.path.abspath(path))
    cfg["matrix"] = os.path.join(base, cfg.get("output_dir", "output"),
                                 cfg.get("matrix_filename", "signal-matrix.xlsx"))
    return cfg


def _rows(ws, first_data_row=4):
    return [r for r in ws.iter_rows(min_row=first_data_row, values_only=True) if r and r[1]]


# People write the Customers cell in whichever shape feels natural: "Acme (2)",
# "Acme: 2", "Acme - 2". Accept all three rather than silently scoring zero.
_ENTRY = re.compile(r"([^,;\n]+?)\s*(?:\((\d)\)|[:\-–]\s*(\d))\s*(?=,|;|\n|$)")


def _entries(cell):
    for m in _ENTRY.finditer(str(cell or "")):
        name = m.group(1).strip()
        value = m.group(2) or m.group(3)
        if name and value:
            yield name, int(value)


def importance_sum(cell):
    return sum(v for _, v in _entries(cell))


def accounts(cell):
    return {name for name, _ in _entries(cell)}


def has_consequence(cell):
    return any(v >= 2 for _, v in _entries(cell))


def compute(matrix_path, minutes_per_call, scope="week", today=None):
    wb = load_workbook(matrix_path)
    signals = _rows(wb["Prioritized Signals"])
    ledger = [r for r in wb["Processed calls"].iter_rows(min_row=4, values_only=True) if r and r[0]]

    today = today or dt.date.today()
    if scope == "week":
        start = today - dt.timedelta(days=today.weekday())
    elif scope == "run":
        start = today
    else:
        start = dt.date.min

    def as_date(v):
        if isinstance(v, dt.datetime):
            return v.date()
        if isinstance(v, dt.date):
            return v
        try:
            return dt.date.fromisoformat(str(v)[:10])
        except ValueError:
            return None

    in_scope = [r for r in ledger if (d := as_date(r[0])) and d >= start]
    by_type = collections.Counter(str(r[5] or "Unclassified") for r in in_scope)

    # Time saved is measured, not assumed: it is the real length of the calls the
    # system read for you. Duration lives in the ledger (column G). Only calls whose
    # length was never recorded fall back to the configured estimate, and the count
    # of those is reported so the number stays auditable.
    def minutes(row):
        raw = row[6] if len(row) > 6 else None
        try:
            return float(raw), True
        except (TypeError, ValueError):
            return float(minutes_per_call), False

    # A three-minute "call" is a no-show or a cancellation. Counting it as time you
    # were spared is the kind of small dishonesty that makes a whole number suspect.
    def real(rows):
        kept, dropped = [], 0
        for r in rows:
            v, measured = minutes(r)
            if measured and v < MIN_REAL_CALL_MINUTES:
                dropped += 1
                continue
            kept.append((v, measured))
        return kept, dropped

    scoped_minutes, noshows = real(in_scope)
    all_minutes, noshows_all = real(ledger)
    measured_scope = sum(1 for _, m in scoped_minutes if m)
    measured_all = sum(1 for _, m in all_minutes if m)
    mins_scope = sum(v for v, _ in scoped_minutes)
    mins_all = sum(v for v, _ in all_minutes)

    ranked = [r for r in signals if not str(r[4] or "").startswith("Out of scope")]
    all_accounts = set()
    with_consequence = 0
    for r in ranked:
        all_accounts |= accounts(r[5])
        if has_consequence(r[5]):
            with_consequence += 1

    no_home = sum(1 for r in ranked
                  if str(r[4] or "").startswith(("Not planned", "Deferred", "Unclear")))
    new_this_scope = sum(1 for r in ranked if (d := as_date(r[10])) and d >= start)
    touched = sum(1 for r in ranked if (d := as_date(r[11])) and d >= start)

    calls = len(scoped_minutes)
    pct = lambda n, d: round(100 * n / d) if d else 0
    return {
        "scope": scope,
        "since": start.isoformat() if start != dt.date.min else "all time",
        "calls": calls,
        "noshows": noshows,
        "prospect": by_type.get("Prospect", 0),
        "implementation": by_type.get("Implementation", 0),
        "customer": by_type.get("Customer", 0),
        "hours_saved": round(mins_scope / 60, 1),
        "minutes_total": round(mins_scope),
        "avg_call_minutes": round(mins_scope / calls) if calls else 0,
        "measured": measured_scope,
        "estimated": calls - measured_scope,
        "minutes_per_call": minutes_per_call,
        "signals_new": new_this_scope,
        "signals_merged": max(touched - new_this_scope, 0),
        "signals_tracked": len(ranked),
        "accounts": len(all_accounts),
        "pct_consequence": pct(with_consequence, len(ranked)),
        "pct_no_roadmap_home": pct(no_home, len(ranked)),
        "calls_all_time": len(all_minutes),
        "hours_saved_all_time": round(mins_all / 60, 1),
        "measured_all": measured_all,
        # Rate is measured over the period the calls actually span, not since the
        # first call — otherwise a backfill of one busy week looks like a slow year.
        "weeks_elapsed": max(
            ((max(_dates) - min(_dates)).days + 1) / 7 if (_dates := [d for r in ledger if (d := as_date(r[0]))]) else 1,
            1),
        "weeks_to_year_end": max((dt.date(today.year, 12, 31) - today).days / 7, 0),
        "today": today.isoformat(),
        "top": sorted(
            ((str(r[1]), r[7] or 0, str(r[3] or "")) for r in ranked),
            key=lambda x: -(x[1] if isinstance(x[1], (int, float)) else 0))[:3],
    }


def render(m, company):
    top = "\n".join(f"  {i}. {name} — priority {p} ({aud})"
                    for i, (name, p, aud) in enumerate(m["top"], 1))
    per_week = m["hours_saved_all_time"] / m["weeks_elapsed"] if m["weeks_elapsed"] else 0
    calls_per_week = m["calls_all_time"] / m["weeks_elapsed"] if m["weeks_elapsed"] else 0
    to_year_end = per_week * m["weeks_to_year_end"]
    noshow_note = (f"; {m['noshows']} no-show(s) under {MIN_REAL_CALL_MINUTES} min excluded"
                   if m["noshows"] else "")
    caveat = ("every call measured" if not m["estimated"]
              else f"{m['measured']} measured, {m['estimated']} estimated at {m['minutes_per_call']} min "
                   f"(length not recorded)")
    return f"""{company} · product discovery — by the numbers ({m['scope']}, since {m['since']})

Coverage
  Calls processed:        {m['calls']}  ({m['prospect']} prospect · {m['implementation']} implementation · {m['customer']} customer)
  All time:               {m['calls_all_time']} calls

Time you did not spend listening
  This {m['scope']}:      {m['hours_saved']} h  ({m['minutes_total']} min of talk time, average call {m['avg_call_minutes']} min)
  All time:               {m['hours_saved_all_time']} h
  Basis:                  {caveat}{noshow_note}

  In working days:        {m['hours_saved']/8:.1f} d this {m['scope']} ({m['hours_saved_all_time']/8:.1f} d all time), at 8 h a day

Run-rate ({calls_per_week:.1f} calls/week, {per_week:.1f} h/week)
  A FULL YEAR at this rate:   {per_week*52:.0f} h = {per_week*52/8:.1f} working days
  Rest of {m['today'][:4]} ({m['weeks_to_year_end']:.0f} weeks left): {to_year_end:.0f} h = {to_year_end/8:.1f} working days
  Total by 31 Dec {m['today'][:4]}:      {to_year_end + m['hours_saved_all_time']:.0f} h = {(to_year_end + m['hours_saved_all_time'])/8:.1f} working days

Backlog
  Signals tracked:        {m['signals_tracked']}
  New / merged:           {m['signals_new']} new · {m['signals_merged']} merged into existing rows
  Accounts represented:   {m['accounts']}

Evidence quality
  Priority driven by real problems: {m['pct_consequence']}% of ranked signals have at least one
  customer who attached a consequence (deal condition, success criterion, quantified pain).
  Demand with no roadmap home:      {m['pct_no_roadmap_home']}%

Top by evidence
{top}
"""


def write_row(matrix_path, m):
    wb = load_workbook(matrix_path)
    ws = wb["Run metrics"]
    ws.append([dt.date.today().isoformat(), m["scope"], m["calls"], m["prospect"],
               m["implementation"], m["customer"], m["hours_saved"], m["signals_new"],
               m["signals_merged"], m["signals_tracked"], m["accounts"],
               f"{m['pct_consequence']}%", f"{m['pct_no_roadmap_home']}%", ""])
    wb.save(matrix_path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default="config.yml")
    ap.add_argument("--matrix", default=None)
    ap.add_argument("--scope", choices=["run", "week", "all"], default="week")
    ap.add_argument("--write", action="store_true", help="append a row to the Run metrics sheet")
    ap.add_argument("--markdown", default=None, help="write the summary to a file")
    args = ap.parse_args()

    cfg = read_config(args.config)
    matrix = args.matrix or cfg["matrix"]
    if not os.path.exists(matrix):
        print(f"no matrix at {matrix} — run tools/create_matrix.py first")
        return 1

    m = compute(matrix, cfg["minutes_per_call_review"], args.scope)
    text = render(m, cfg["company"])
    print(text)
    if args.write:
        write_row(matrix, m)
        print("appended to 'Run metrics'")
    if args.markdown:
        with open(args.markdown, "w", encoding="utf-8") as fh:
            fh.write("```\n" + text + "```\n")
        print(f"wrote {args.markdown}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
