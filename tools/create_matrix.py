#!/usr/bin/env python3
"""Create the blank signal matrix.

Four sheets:
  Prioritized Signals — one row per signal, cumulative, never rebuilt
  Cross-call themes   — patterns seen across 2+ calls
  Processed calls     — the idempotency ledger (what makes runs self-healing)
  Run metrics         — what the system reports about itself

Layout convention on every sheet: row 1 title, row 2 legend, row 3 headers, data from row 4.

Usage:  python3 tools/create_matrix.py [--config config.yml] [--force]
"""
import argparse
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_BG = "1A1A1A"
GREY = "777777"

SIGNAL_HEADERS = [
    ("Rank", 6), ("Signal", 32), ("Type", 27), ("Audience", 23),
    ("Roadmap status", 27), ("Customers (importance 1-2)", 26), ("Mentions", 11),
    ("Priority", 11), ("Pain point / request", 58), ("Customer quotes", 60),
    ("First seen", 12), ("Last mentioned", 13), ("Owner", 16), ("Sources", 40),
]
THEME_HEADERS = [("Date", 12), ("Theme", 55), ("Calls", 8), ("Customers", 40)]
LEDGER_HEADERS = [
    ("Call date", 12), ("Call title / customer", 60), ("Ran by", 18),
    ("Call id / link", 45), ("Processed on", 13), ("Call type", 14),
    ("Duration (min)", 14),
]
METRIC_HEADERS = [
    ("Date", 12), ("Scope", 10), ("Calls processed", 15), ("Prospect", 10),
    ("Implementation", 14), ("Customer", 10), ("Listening time saved (h)", 24),
    ("Signals new", 12), ("Signals merged", 14), ("Signals tracked", 15),
    ("Accounts represented", 20), ("% with consequence", 18),
    ("% no roadmap home", 18), ("Open questions", 14),
]

SIGNAL_LEGEND = (
    "Priority = (sum of customer importances) x mentions — pure evidence, nothing estimated. "
    "Importance per customer (a customer is only listed if they raised it): "
    "2 = a CONSEQUENCE is attached (deal condition, pilot success criterion, hard constraint, "
    "quantified pain, renewal risk) · 1 = interest without consequence (\"would be nice\"). "
    "Effort and feasibility are discussed with engineering, not stored here. "
    "One row per signal — rows are UPDATED, never duplicated. "
    "Rows marked 'Out of scope' were closed by a human and sit at the bottom, unranked."
)
METRIC_LEGEND = (
    "Written by tools/metrics.py from the matrix and the ledger — nothing tracked by hand. "
    "'Listening time saved' is MEASURED: it is the summed real duration of the calls the system read "
    "for you, taken from the ledger. Only calls whose length was never recorded fall back to the "
    "minutes_per_call_review estimate in config, and reports say how many. "
    "'% with consequence' = share of ranked signals where at least one customer attached a "
    "consequence (importance 2) — the check that ranking follows real customer problems."
)


def _sheet(wb, title, headers, legend, sheet_title, freeze=None, autofilter=False, first=False):
    ws = wb.active if first else wb.create_sheet()
    ws.title = title
    ws["A1"] = sheet_title
    ws["A1"].font = Font(bold=True, size=13, color="1A1A1A")
    ws["A2"] = legend
    ws["A2"].font = Font(italic=True, size=9, color=GREY)
    for col, (name, width) in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col, value=name)
        cell.font = Font(bold=True, size=10, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = width
    if freeze:
        ws.freeze_panes = freeze
    if autofilter:
        last = ws.cell(row=3, column=len(headers)).column_letter
        ws.auto_filter.ref = f"A3:{last}3"
    return ws


def build(company: str) -> Workbook:
    wb = Workbook()
    _sheet(wb, "Prioritized Signals", SIGNAL_HEADERS, SIGNAL_LEGEND,
           f"{company} · Signal Prioritization Matrix — cumulative across all calls",
           freeze="C4", autofilter=True, first=True)
    _sheet(wb, "Cross-call themes", THEME_HEADERS,
           "Patterns that showed up in 2 or more calls. Appended, never overwritten.",
           "Cross-call themes")
    _sheet(wb, "Processed calls", LEDGER_HEADERS,
           "The idempotency ledger. Every processed call is logged here so a call is never counted "
           "twice and a missed run heals itself. Do not delete rows.",
           "Processed calls")
    _sheet(wb, "Run metrics", METRIC_HEADERS, METRIC_LEGEND, "Run metrics")
    return wb


def read_company(config_path: str) -> str:
    """Tiny config reader — avoids a PyYAML dependency for one string."""
    if not os.path.exists(config_path):
        return "Your company"
    in_company = False
    for line in open(config_path, encoding="utf-8"):
        stripped = line.strip()
        if stripped.startswith("company:"):
            in_company = True
            continue
        if in_company:
            if stripped.startswith("name:"):
                return stripped.split(":", 1)[1].strip().strip('"\'')
            if line and not line[0].isspace():
                break
    return "Your company"


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--config", default="config.yml")
    ap.add_argument("--out", default=None, help="output path (default: ./output/signal-matrix.xlsx)")
    ap.add_argument("--force", action="store_true", help="overwrite an existing matrix")
    args = ap.parse_args()

    out = args.out or os.path.join("output", "signal-matrix.xlsx")
    if os.path.exists(out) and not args.force:
        print(f"refusing to overwrite existing matrix: {out}\n"
              f"(this file is the accumulated evidence base — pass --force only if you mean it)")
        return 1

    os.makedirs(os.path.dirname(out) or ".", exist_ok=True)
    build(read_company(args.config)).save(out)
    print(f"created {out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
