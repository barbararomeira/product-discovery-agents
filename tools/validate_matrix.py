#!/usr/bin/env python3
"""Check the matrix the agent just wrote, before anyone reads it.

Why this exists rather than tests over the merge itself: the merge *is* the judgement. Deciding
that two customers described the same underlying need is the one thing here a model does that
rules cannot, and pretending otherwise by unit-testing a merge function would mean writing a
merge function that does not — and should not — exist.

What you can do is check the model's output deterministically. Every rule below is arithmetic or
set membership, so it is the same answer every time, and a violation means the agent broke a rule
the prompt states plainly. That is a real class of failure: a run that mis-parses one cell, or
quietly renumbers a rank, produces a matrix that looks finished and is wrong.

Exit codes: 0 clean, 1 violations found, 2 could not read the matrix at all.

Usage:
    validate_matrix.py <matrix.xlsx>
    validate_matrix.py <matrix.xlsx> --save-manual manual.json   # before a run
    validate_matrix.py <matrix.xlsx> --check-manual manual.json  # after a run
"""
from __future__ import annotations

import argparse
import json
import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from metrics import _entries  # noqa: E402  — one parser, not two

SHEET = "Prioritized Signals"
HEADER_ROW = 3
FIRST_DATA_ROW = 4

# Columns the agent must never overwrite: they are the PM's, and an automated merge that "tidies"
# them is indistinguishable from data loss. Owner is the only one today; add to this tuple rather
# than scattering the rule.
MANUAL_COLUMNS = ("Owner",)


def _load(path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET not in wb.sheetnames:
        raise KeyError(f"sheet {SHEET!r} not found (have: {', '.join(wb.sheetnames)})")
    ws = wb[SHEET]
    header = [str(c.value).strip() if c.value else "" for c in ws[HEADER_ROW]]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=FIRST_DATA_ROW, values_only=True), FIRST_DATA_ROW):
        if not any(row):
            continue
        rows.append((i, dict(zip(header, row))))
    return header, rows


def _int(value):
    """Spreadsheet numbers arrive as int, float or str depending on how they were written."""
    if value is None or str(value).strip() == "":
        return None
    try:
        return int(float(str(value).strip()))
    except ValueError:
        return None


def validate(header, rows) -> list[str]:
    problems = []
    ranks = []

    for line, r in rows:
        signal = str(r.get("Signal") or "?")[:40]
        where = f"row {line} ({signal})"

        entries = list(_entries(r.get("Customers (importance 1-2)")))
        if not entries:
            problems.append(f"{where}: no customer carries an importance — the row has no evidence")

        # 1. importance is 1 or 2, never anything else. A 3 means the scale was invented.
        for name, value in entries:
            if value not in (1, 2):
                problems.append(f"{where}: importance {value} for {name!r} — the scale is 1 or 2 only")

        # 2. one account, one importance. A repeated account inflates the sum and the priority.
        names = [n for n, _ in entries]
        dupes = {n for n in names if names.count(n) > 1}
        if dupes:
            problems.append(f"{where}: {', '.join(sorted(dupes))} listed more than once — "
                            "customers are counted by account, not by call")

        mentions = _int(r.get("Mentions"))
        if mentions is None or mentions < 1:
            problems.append(f"{where}: Mentions is {r.get('Mentions')!r} — expected a positive whole number")

        # 3. the priority formula, recomputed independently of whatever the agent wrote.
        priority = _int(r.get("Priority"))
        if entries and mentions and mentions >= 1:
            expected = sum(v for _, v in entries) * mentions
            if priority != expected:
                problems.append(f"{where}: Priority is {priority}, expected {expected} "
                                f"(importance sum {sum(v for _, v in entries)} × {mentions} mentions)")

        # 4. every claim traces to a call. A row with no source cannot be checked by a human.
        if not str(r.get("Sources") or "").strip():
            problems.append(f"{where}: no source call cited")

        rank = _int(r.get("Rank"))
        if rank is None:
            problems.append(f"{where}: Rank is missing")
        else:
            ranks.append((rank, priority if priority is not None else -1, line))

    # 5. ranks are 1..n, each used once, ordered by descending priority. A silent renumber is the
    # failure this catches: the numbers stay plausible and the order stops meaning anything.
    if ranks:
        seen = sorted(r for r, _, _ in ranks)
        if seen != list(range(1, len(seen) + 1)):
            problems.append(f"Rank is not 1..{len(seen)} exactly once (got {seen[:12]}…)")
        by_rank = sorted(ranks)
        for (r1, p1, l1), (r2, p2, l2) in zip(by_rank, by_rank[1:]):
            if p1 < p2:
                problems.append(f"rank {r1} (priority {p1}, row {l1}) sits above "
                                f"rank {r2} (priority {p2}, row {l2}) — ranking is not by priority")
                break
    return problems


def manual_snapshot(header, rows) -> dict:
    cols = [c for c in MANUAL_COLUMNS if c in header]
    return {str(r.get("Signal")): {c: r.get(c) for c in cols} for _, r in rows}


def compare_manual(before: dict, header, rows) -> list[str]:
    """A signal the PM has annotated may be re-ranked and re-evidenced, never re-owned."""
    now = manual_snapshot(header, rows)
    problems = []
    for signal, fields in before.items():
        if signal not in now:
            continue  # rows can legitimately be merged away; that is the agent's job
        for col, was in fields.items():
            became = now[signal].get(col)
            if was != became and was not in (None, ""):
                problems.append(f"{signal!r}: {col} was {was!r} and is now {became!r} — "
                                "manual fields belong to the PM and are never overwritten")
    return problems


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__.splitlines()[0])
    ap.add_argument("matrix")
    ap.add_argument("--save-manual", metavar="FILE", help="write a snapshot of PM-owned fields")
    ap.add_argument("--check-manual", metavar="FILE", help="compare PM-owned fields against a snapshot")
    args = ap.parse_args()

    try:
        header, rows = _load(args.matrix)
    except Exception as exc:                                   # unreadable, missing, wrong shape
        print(f"VALIDATE: cannot read {args.matrix}: {exc}", file=sys.stderr)
        return 2

    if args.save_manual:
        pathlib.Path(args.save_manual).write_text(
            json.dumps(manual_snapshot(header, rows), indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"VALIDATE: manual-field snapshot written for {len(rows)} rows")
        return 0

    problems = validate(header, rows)
    if args.check_manual and pathlib.Path(args.check_manual).exists():
        before = json.loads(pathlib.Path(args.check_manual).read_text(encoding="utf-8"))
        problems += compare_manual(before, header, rows)

    if problems:
        print(f"VALIDATE: {len(problems)} problem(s) in {args.matrix}", file=sys.stderr)
        for p in problems:
            print(f"  - {p}", file=sys.stderr)
        return 1
    print(f"VALIDATE: {len(rows)} rows clean")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
